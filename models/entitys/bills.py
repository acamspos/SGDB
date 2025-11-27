from pydantic import BaseModel
from typing import Optional
from assets.db.db_connection import DB
from datetime import datetime, timedelta
from models.entitys.activity import Activity
import pathlib
from win32com import client
import openpyxl
from openpyxl.drawing.image import Image
import shutil
from datetime import datetime
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
from tkinter import messagebox
from mysql.connector import Error

class Bill(BaseModel):
   
    code: int| None = None
    description: str
    client: str = ""
    creationDate: datetime = datetime.today().strftime('%Y-%m-%d')
    expirationDate: datetime | None = None
    currency: int = 1
    purchaseOrder: str | None = None
    documentCondition: str = 'CREDITO'
    payment_status: int = 1
    budget_code: int |  None = None
    sub_total: float = 0
    iva: float = 0
    total_amount: float = 0
    totalUSD: float
    debtUSD: float = 0 
    totalPaidUSD: float = 0
    exchange_rate: float = 0
    documentState: int = 1
    notes: str | None = None
    creationUser: int | None = None
    processingUser: int | None = None


    def create(self, items=None):
        dict_data = self.model_dump()
        schema = tuple(dict_data.keys())[1:]
        data = list(dict_data.values())[1:]
        self.code = DB.createBillDB(schema=schema, data=data)
        if not items:
            items =  Item.findItemsFromBudget(self.budget_code)
            for x in range(len(items)):
                itemCurrency = items[x]['currency']
                if itemCurrency != self.currency:
                    product_rate = 1
                    document_rate = float(self.exchange_rate)

                    if self.currency == 2:
                        product_rate, document_rate = document_rate, product_rate

                    new_price = round(float(items[x]['price']),2) if self.currency == itemCurrency else round(float(items[x]['price']) * document_rate/product_rate, 2)
                    new_total_price = new_price * items[x]['quantity'] 
                    items[x]['price']=new_price
                    items[x]['total_price']=new_total_price
                    items[x]['currency']=self.currency

        Item.create_list(items, self.code)

    def openDecorator(func):
        def nueva_funcion(self,*args,**kwargs):
            try:
                DB.DB.connect()
                data = func(self,*args,**kwargs)
              
                if data:
                    return data
            except Error as ex:
                messagebox.showerror('Error', f"Ha ocurrido un error en la base de datos: {ex}")
            finally:
                DB.DB.close()
            
        return nueva_funcion
       
    def findItems(self) -> list:
        records = Item.findAllItems(self.code)
        return [Item(**data) for data in records]
    
    def update(self):
        dict_data = self.model_dump()
        schema = tuple(dict_data.keys())[1:]
        data = list(dict_data.values())[1:]
        data.append(self.code)
        sentence = ''
        for key in schema:
            sentence+=f"{key}=%s,"
        sentence = sentence[:-1]
        DB.updateBillDB(fields=sentence, data=data)

    @openDecorator
    def get_documentState(self):
        aux = DB.DB.cursor(buffered=True)
        sql = 'SELECT description FROM documentstate WHERE id=%s'
        aux.execute(sql, (self.documentState,))
        currencyD = aux.fetchone()[0]
        return currencyD
    @openDecorator
    def get_documentStatus(self):
        aux = DB.DB.cursor(buffered=True)
        sql = 'SELECT description FROM paymentstatus WHERE id=%s'
        aux.execute(sql, (self.payment_status,))
        currencyD = aux.fetchone()[0]
        return currencyD

    
    def rejectBill(self,user):
        DB.rejectBill(self.code, user)

    def deleteBill(self):
        DB.deleteOneBillDB(self.code)

    def update_reject_status(self):
        DB.updateRejectBill(self.code)

    @classmethod
    def findAllBills(self, value='',condition_field = 'b.code', codition_operator = 'LIKE', filter_status = None):
        data = DB.findAllBillsDB(value, condition_field, codition_operator, filter_status)
        return data
    
    @classmethod
    def getNextBillcode(self):
        return DB.getNextBillCode2()


    @classmethod
    def findRejectsBills(self, ):
        data = DB.findRejectBillsDB()
        return data
    
    @classmethod
    def findOneBill(self, value=''):
        data = DB.findOneBillDB(value)
        return Bill(**data)

    @openDecorator
    def get_company(self):
        aux = DB.DB.cursor(buffered=True)
        sql = 'SELECT name FROM clients WHERE rif=%s'
        aux.execute(sql, (self.client,))
        company = aux.fetchone()[0]
        return company
    
    @openDecorator
    def get_company_address(self):
        aux = DB.DB.cursor(buffered=True)
        sql = 'SELECT address FROM clients WHERE rif=%s'
        aux.execute(sql, (self.client,))
        company = aux.fetchone()[0]
        return company

    @openDecorator
    def get_currency(self):
        aux = DB.DB.cursor(buffered=True)
        sql = 'SELECT description FROM currency WHERE id=%s'
        aux.execute(sql, (self.currency,))
        currencyD = aux.fetchone()[0]
        return currencyD

    @openDecorator
    def get_currency_icon(self):
        aux = DB.DB.cursor(buffered=True)
        sql = 'SELECT icon FROM currency WHERE id=%s'
        aux.execute(sql, (self.currency,))
        currencyD = aux.fetchone()[0]
        return currencyD
    

    @classmethod
    def update_document_states(self):
        return DB.updateBillsToExpire()
    
    @classmethod
    def get_document_to_expirate(self):
        return [Bill(**doc) for doc in DB.findBillsToExpira()]

    @classmethod
    def totalCountBills(self):
        return DB.count_total_bills_DB()


    @classmethod
    @openDecorator
    def countBills(self):
       
            aux = DB.DB.cursor(buffered=True)
            month = datetime.today().month
            year = datetime.today().year
            sql = f"""
            SELECT 
                DATE(creationDate) AS fecha,
                COUNT(*) AS cantidad_registros
            FROM 
                bills
            WHERE 
                MONTH(creationDate) = {month} AND YEAR(creationDate) = {year}
            GROUP BY 
                DATE(creationDate);"""
            aux.execute(sql,)
            currencyD = aux.fetchall()
  
            return currencyD
      

    
    @classmethod
    @openDecorator
    def lastFive(self, num_docs = 5):
        #SELECT * FROM bills ORDER BY id DESC LIMIT 5;last man standing livingstones

        aux = DB.DB.cursor(buffered=True, dictionary=True)
        month = datetime.today().month
        year = datetime.today().year
        sql = f"""SELECT * FROM bills ORDER BY code DESC LIMIT %s"""
        aux.execute(sql,(num_docs,))
        currencyD = aux.fetchall()
        
        return currencyD
    
    def create_pdf(self, path_to_save, parent=None):    
       
        template = str(pathlib.Path(__file__).parent.parent.parent) + "\\templates\\bill2.xlsx"
        template_file = openpyxl.load_workbook(template)
        sheet = template_file['template']
        sheet = template_file.active
        
        nueva = sheet
        nueva.title = f'Factura {self.code}'
        
        nueva['L1'] = '0'*(6-len(str(self.code)))+str(self.code)
        nueva['B2'] = self.get_company()
        nueva['B3'] = self.client
        nueva['B4'] = self.get_company_address()
        nueva['J2'] = self.creationDate
        nueva['J3'] = self.expirationDate
        nueva['J4'] = self.purchaseOrder
        nueva['A23'] = f"""NOTA: ESTA FACTURA TIENE BASE IMPONIBLE DE {f"{self.totalUSD} $" if self.currency==1 else f"Bs. {round(self.totalUSD*self.exchange_rate,2)}"}  CALCULADOS A LA TASA DEL BCV DEL 
DIA {self.creationDate.strftime('%d/%m/%Y')} A RAZON DE BSD {self.exchange_rate}"""
        nueva['A35'] = self.notes

        nueva['K32'] = f"SUB - TOTAL {self.get_currency_icon()}"
        nueva['K34'] = f"TOTAL {self.get_currency_icon()}"

        border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        items = [Item(**ite) for ite in Item.findAllItems(self.code)]
        num = 8
        for index, item in enumerate(items):
            if not item.position:
                position = num+index
                nueva[f'A{position}'] = '0'*(3-len(str(index+1)))+str(index+1)
                nueva[f'B{position}'] = item.itemDescription
                nueva[f'I{position}'] = item.quantity
                nueva.merge_cells(f'J{position}:K{position}')
                nueva[f'J{position}'] = item.price
                nueva[f'L{position}'] = item.total_price
                for row in nueva[f'A{position}:L{position}']:
                    for cell in row:
                        cell.border = border_style
            else:
                 position = num+index*2
                 nueva[f'A{position}'] = 'POS. '+'0'*(5-len(str(item.position)))+str(item.position)
                 nueva.merge_cells(f'A{position}:A{position+1}')
                 nueva[f'B{position}'] = item.itemDescription
                 nueva[f'B{position+1}'] = f'DOCUMENTO DE ACEPTACION NUMERO:{item.acceptance_code}'
                 nueva[f'I{position}'] = item.quantity
                 nueva.merge_cells(f'I{position}:I{position+1}')
                 nueva.merge_cells(f'J{position}:K{position+1}')
                 nueva[f'J{position}'] = item.price
                 
                 nueva[f'L{position}'] = item.total_price
                 nueva.merge_cells(f'L{position}:L{position+1}')
                 num+=1

        border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for row in sheet[f'A{8}:L{num}']:
            for cell in row:
                cell.border = border_style
    
        output_template = str(pathlib.Path(__file__).parent.parent.parent) + "\\templates\\BILLS.xlsx"
        template_file.save(output_template)
        template_file.close()


        try:
            code = '0'*(7-len(str(self.code)))+str(self.code)
            path = str(path_to_save).replace('/','\\') + f'\\Factura_{code}.pdf'
     
            excel = client.Dispatch("Excel.Application")

            sheets = excel.Workbooks.Open(output_template)
            work_sheets = sheets.Worksheets[0]
            # Converting into PDF File
            work_sheets.ExportAsFixedFormat(0, path)
        except:
            excel.quit()
            return None
        finally:
            sheets.Close()
            excel.quit()
            import os
            os.remove("C:\\Users\\Agustin Campos\\Desktop\\SGDB\\templates\\BILLS.xlsx")
            messagebox.showinfo('Factura','PDF generado exitosamente.', parent=parent)
            return(path)
    

class Item(BaseModel):
   
    code: int | None = None
    itemType: int
    itemId: str
    position: int | None = None
    acceptance_code: str  | None = None
    itemDescription: str
    cost: float
    quantity: int
    price: float
    total_price: float
    totalUSD: float
    currency: int
    date: datetime | str

    def openDecorator(func):
        def nueva_funcion(self):
            try:
                DB.DB.connect()
                func(self)
            except Error as ex:
                messagebox.showerror('Error', f"Ha ocurrido un error en la base de datos: {ex}")
            finally:
                DB.DB.close()
            
        return nueva_funcion

    def create(self):
        dict_data = self.model_dump()
        schema = tuple(dict_data.keys())
        data = list(dict_data.values())
        DB.createItemBillsDB(schema=schema, data=data)

    @openDecorator
    def get_type(self):
        aux = DB.DB.cursor(buffered=True)
        sql = 'SELECT description FROM itemType WHERE id=%s'
        aux.execute(sql, (self.itemType,))
        typeD = aux.fetchone()[0]
        return typeD

    @classmethod
    def create_list(self, items, bill):
        for item in items:
            item['code'] = bill
            Item(**item).create()


    @openDecorator
    @classmethod
    def set_documentCode(self,data):
        aux = DB.DB.cursor(buffered=True)
        sql = 'UPDATE itemsBills SET position=%s,acceptance_code=%s WHERE itemid=%s AND code=%s'
        aux.executemany(sql, data)
        DB.DB.commit()
        aux.close()



    @classmethod
    def findAllItems(self, bill):
        data = DB.findABillsitemsDB(bill)
        return data
    

    @classmethod
    def findItemsFromBudget(self, budget,):
        data = DB.findABudgetitemsDB(budget)

        return data
    
    

    
    @classmethod
    def deleteItems(self, budget_code):
        DB.deleteAllItemsByBudget(budget_code)

    @classmethod
    def normalQuantity(self, budget_code, itemId):
        data = DB.returnQuantityOneItemsDB(budget_code,itemId)
        return data
    


