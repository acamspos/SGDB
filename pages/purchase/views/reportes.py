import ttkbootstrap as ttb
from tkinter import ttk
from assets.globals import GUI_COLORS, GUI_FONT, IMG_PATH

from models.entitys.product import  Product
from pages.providers.providers import ProviderModule
from tkinter import messagebox
from assets.styles.styles import SGDB_Style
from customtkinter import CTkFrame
from PIL import Image, ImageTk
from components.buttons import ButtonImage
from assets.utils import resize_icon, resize_image
import tkinter as tk
from assets.db.db_connection import DB
from tkinter import filedialog
import os
from datetime import datetime
from sqlalchemy import create_engine
import pandas as pd
from openpyxl.styles import Border, Side
from openpyxl.utils.cell import get_column_interval
import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False

def on_combobox_change(event, var, dictionary, combobox):
    ID = dictionary[combobox.get()]
    var.set(ID)

    


def on_combobox_change(event, var, dictionary, combobox):
    ID = dictionary[combobox.get()]
    var.set(ID)




class Report(ttb.Toplevel):
    def __init__(self, master=None, window_type = 'create', title = '', product = None):
        super().__init__(master, toolwindow=True)
        self.valid_character = '012356789.-'
        self.focus()

        # SYSTEM_WIDTH = self.winfo_screenwidth()
        # SYSTEM_HEIGHT = self.winfo_screenheight()

        # pwidth = (SYSTEM_WIDTH-1158)//2
        # pheight = (SYSTEM_HEIGHT-794)//2

        # self.geometry(str(1158)+"x"+str(794)+"+"+str(pwidth)+"+"+str(pheight-60))
        # self.minsize(width=1158,height=794)

        
        self.window_title = title
        self.window_type = window_type
        self.__PROVIDER_RIF = ttb.StringVar()
        self.__state = ttb.StringVar()
        self.__status = ttb.StringVar()
        self.fileNameVar = ttb.StringVar(value=f"Reporte_de_Compras_{datetime.today().strftime('%d-%m-%Y')}")

        ######### MODAL WINDOW CONFIG #########
   
        self.title(title)
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)
        self.transient()
        self.grab_set()
        self.config(background="#D9D9D9")
       
        self.__createWidgets()
        self.place_window_center()

    


    
    def __createWidgets(self):
        


        FGCOLOR = 'white'

        contentFrame = CTkFrame(self, fg_color='white',border_width=1,  border_color='#CFCFCF')
        contentFrame.grid(row = 1, column = 0, sticky = 'nsew', padx=10, pady=10)
        contentFrame.columnconfigure(0, weight=1)


        self.micon = resize_icon(Image.open(f"{IMG_PATH}/reporticon.png"))


        subtitle = CTkFrame(contentFrame, fg_color=GUI_COLORS['primary'],)
        subtitle.grid(row = 0, column = 0, sticky = 'nsew', padx=12, pady=(8,0))
        
        ttb.Label(subtitle, image=self.micon, background=GUI_COLORS['primary'], ).grid(row=0, column=0, rowspan=2, sticky='nsew', padx=(8), pady=8)

        ttb.Label(subtitle, text='Reportes', background=GUI_COLORS['primary'], font=(GUI_FONT,13,'bold'), foreground='#fff', anchor='sw').grid(row=0, column=1, sticky='nsew', padx=(4,8), pady=(8,0))

        ttb.Label(subtitle, text=f'Inicio / {self.window_title} / Reportes', background=GUI_COLORS['primary'], font=(GUI_FONT,9), foreground='#fff', anchor='nw').grid(row=1, column=1, sticky='nsew', padx=(4,8), pady=(0,8))



        ttb.Frame(contentFrame, bootstyle='dark').grid(row=1, column=0, sticky='nsew', padx=12, pady=5)

    
            ########### Description Section ###########

        mainInforFrame = tk.Frame(contentFrame)
        mainInforFrame.config(background=FGCOLOR)
        mainInforFrame.grid(row=2, column=0, sticky='nsew', padx=20, pady=10)
        mainInforFrame.columnconfigure(0, weight=1)
        mainInforFrame.columnconfigure(1, weight=1)
 


        code_label = ttb.Label(mainInforFrame, 
                                      anchor='w', 
                                      text='Opciones de Filtrado:', 
                                      bootstyle='primary', 
                                      background=FGCOLOR,
                                      font=('arial',11,'bold'))
        code_label.grid(row=0, column=0, padx=(8,0), pady=(2,10), ipadx=8, ipady=8, sticky='nsew')



        
        deliverytimelabel = ttb.Label(mainInforFrame, 
                                   anchor='w', 
                                   text='Por Status de Pago', 
                                   background=FGCOLOR,
                                   font=(GUI_FONT,11,'bold'))
        deliverytimelabel.grid(row=1, column=0, padx=5, pady=(4,0), ipadx=8,  ipady=4,sticky='nsew')

        self.STATUS = {
            'Todas':None,
            'Por Pagar': 1,
            'Pagadas': 2
        }
     

        self.statusSelection = ttb.Combobox(mainInforFrame, values=list(self.STATUS.keys()),state='readonly',font=(GUI_FONT,10,'bold'),
                                    style='selectionOnly.TCombobox', textvariable=self.__status)
        self.statusSelection.grid(row=2, column=0, padx=(4,10), pady=(2,0), sticky='nsew')
        self.statusSelection.current(0)
        deliverytimelabel = ttb.Label(mainInforFrame, 
                                   anchor='w', 
                                   text='Por Estado del Documento', 
                                   background=FGCOLOR,
                                   font=(GUI_FONT,11,'bold'))
        deliverytimelabel.grid(row=1, column=1, padx=5, pady=(4,0), ipadx=8,  ipady=4,sticky='nsew')
     
        self.STATES = {
            'Todas':None,
            'En Rango':1,
            'Vencidas':2,
            'Procesadas':3
        }

        self.state_selection = ttb.Combobox(mainInforFrame, values=list(self.STATES.keys()),state='readonly',font=(GUI_FONT,10,'bold'),
                                    style='selectionOnly.TCombobox', textvariable=self.__state)
        self.state_selection.grid(row=2, column=1, padx=(4,10), pady=(2,0), sticky='nsew')
        self.state_selection.current(0)

        deliverytimelabel = ttb.Label(mainInforFrame, 
                                   anchor='w', 
                                   text='Por Cliente', 
                                   background=FGCOLOR,
                                   font=(GUI_FONT,11,'bold'))
        deliverytimelabel.grid(row=1, column=2, padx=5, pady=(4,0), ipadx=8,  ipady=4,sticky='nsew')
     

        self.providerRIFEntry = ttb.Entry(mainInforFrame, state='readonly', justify='center', width=30, textvariable=self.__PROVIDER_RIF)
        self.providerRIFEntry.grid(row=2, column=2, sticky='nsw')

        
        
        self.provider_search_btn = ttb.Button(mainInforFrame, 
                                            command=self.open_provider_selection,
                                            text='...',
                                            bootstyle='dark')
        self.provider_search_btn.grid(row=2, column=3, padx=2, sticky='nsew', pady=1)

        self.provider_search_btn = ttb.Button(mainInforFrame, 
                                            command=self.clean_provider_field,
                                            text='Borrar',
                                            bootstyle='dark')
        self.provider_search_btn.grid(row=2, column=4, padx=2, sticky='nsew', pady=1)
        

        filenameLabel = ttb.Label(mainInforFrame, 
                                   anchor='w', 
                                   text='Nombre del Archivo', 
                                   background=FGCOLOR,
                                   font=(GUI_FONT,11,'bold'))
        filenameLabel.grid(row=3, column=0, padx=5, pady=(4,0), ipadx=8,  ipady=4,sticky='nsew')


        self.filenameEntry = ttb.Entry(mainInforFrame, justify='center', bootstyle='primary', width=30, textvariable=self.fileNameVar)
        self.filenameEntry.grid(row=4, column=0, columnspan=5, ipady=4, sticky='nesw')
        


        ttb.Separator(contentFrame, bootstyle='light').grid(row=3, column=0, sticky='nsew', padx=10)


        buttonss_section_frame = tk.Frame(contentFrame,)
        buttonss_section_frame.configure(background=FGCOLOR)
        buttonss_section_frame.grid(row=4, column=0, pady=(8,8), sticky='nsew', padx=10)
        buttonss_section_frame.anchor('e')
        





        closeimg = Image.open(f"{IMG_PATH}/closen.png")
        self.closeimg = ImageTk.PhotoImage(closeimg.resize(resize_image(20, closeimg.size)))
        closeimgh = Image.open(f"{IMG_PATH}/closeh.png")
        self.closeimgh = ImageTk.PhotoImage(closeimgh.resize(resize_image(20, closeimgh.size)))
        closeimgp = Image.open(f"{IMG_PATH}/closep.png")
        self.closeimgp = ImageTk.PhotoImage(closeimgp.resize(resize_image(20, closeimgp.size)))
        self.closeBTN = ButtonImage(buttonss_section_frame, image=self.closeimg, img_h=self.closeimgh, img_p=self.closeimgp, command=self.destroy, style='flatw.light.TButton', text='CERRAR', compound='center',padding=0)
        self.closeBTN.grid(row=0, column=1, sticky='nsew', pady=2, padx=(0,4))

        
        #self.__form_fields =[self.codeEntry, self.descriptionEntry, self.departmentEntry, self.brandEntry, self.providerEntry, self.taxCombobox, self.measurementCombobox, self.currencyCombobox, self.costEntry, self.directCostEntry, self.indirectCostEntry,
                           #self.priceEntry1, self.priceEntry2, self.priceEntry3, self.profitEntry1, self.profitEntry2, self.profitEntry3, self.mainStockEntry, self.depositEntry1, self.depositEntry2, self.depositEntry3, self.depositEntry4]
        
        #self.__form_variables = [self.code_var, self.description_var, self.department_var, self.brand_var, self.provider_var, self.tax_var, self.measurement_var, self.currency_var, self.cost_var, self.indirect_cost_var, self.direct_cost_var,
            #                     self.price_1_var, self.price_2_var, self.price_3_var, self.profit_1_var, self.profit_2_var, self.profit_3_var, self.stock_var, self.stock_1_var, self.stock_2_var, self.stock_3_var, self.stock_4_var]
        

        creatbtnimg = Image.open(f"{IMG_PATH}/registrar.png")
        self.creatbtnimg = ImageTk.PhotoImage(creatbtnimg.resize(resize_image(20, creatbtnimg.size)))
        creatbtnimgh = Image.open(f"{IMG_PATH}/registrarh.png")
        self.creatbtnimgh = ImageTk.PhotoImage(creatbtnimgh.resize(resize_image(20, creatbtnimgh.size)))
        creatbtnimgp = Image.open(f"{IMG_PATH}/registrarp.png")
        self.creatbtnimgp = ImageTk.PhotoImage(creatbtnimgp.resize(resize_image(20, creatbtnimgp.size)))

        self.createBTN = ButtonImage(buttonss_section_frame, image=self.creatbtnimg, img_h=self.creatbtnimgh, img_p=self.creatbtnimgp, 
                                     style='flatw.light.TButton', text='GENERAR', compound='center',padding=0, command=self.generate_report)
        self.createBTN.grid(row=0, column=3, sticky='nsew', pady=2, padx=(0,4))

    def set_provider_rif(self, provider):
        self.__PROVIDER_RIF.set(provider.rif)
        self.fileNameVar.set(f"{provider.rif}_"+self.fileNameVar.get())

    def clean_provider_field(self):
      
        if self.__PROVIDER_RIF.get():
            self.fileNameVar.set(self.fileNameVar.get().replace(f"{self.__PROVIDER_RIF.get()}_",''))
            self.__PROVIDER_RIF.set('')

    def open_provider_selection(self):
        window = ProviderModule(self,selectionMode=True, callback=self.set_provider_rif)


    def get_sql_report_sentence(self):
        sql = """SELECT p.code as Codigo, p.control as Control, pv.name as Empresa, p.provider as RIF, p.documentCondition as Condicion, p.dateOfIssue as 'Fecha de Emisión',
        p.expirationDate as 'Fecha de Vencimiento', p.registrationDate as 'Fecha de Registro',  c.description as Moneda, p.total as 'Total Principal', p.exchangeRate as 'Tasa de Cambio', p.totalUSD as 'Total USD $', p.debtUSD as 'Deuda USD $', p.totalPaidUSD as 'Total Pagado USD $', 
        ds.description as Estado, ps.description as Status, pt.description as Tipo, p.description FROM purchases p
        INNER JOIN providers pv ON pv.rif = p.provider
        INNER JOIN currency c ON c.id=p.currency
        INNER JOIN documentstate ds ON ds.id = p.documentState
        INNER JOIN paymentstatus ps ON ps.id = p.payment_status
        INNER JOIN billtype pt ON pt.id = p.purchaseType"""

        extra_fields = []

        payment_status = self.STATUS[self.__status.get()]
        if payment_status:
            extra_fields.append(f'payment_status = {payment_status}')
        
        document_state = self.STATES[self.__state.get()]
        if document_state:
            extra_fields.append(f"documentState = {document_state}")

        rif = self.__PROVIDER_RIF.get()
        if rif:
            extra_fields.append(f"provider = '{rif}'")
        
        for index, field in enumerate(extra_fields):
            if index == 0:
                sql+=f' WHERE {field}'
            else:
                sql+=f' AND {field}'
        return sql
    


    def generate_report(self):
            desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
            path = filedialog.askdirectory(initialdir=desktop_path,parent=self)
            if path:
                path += f'/{self.fileNameVar.get()}.xlsx'
                db_connection_str =f'mysql+mysqlconnector://root:2692002Acc.@localhost/billssg_db'
                db_connection = create_engine(db_connection_str)
                sql = self.get_sql_report_sentence()
                df = pd.read_sql(sql, con=db_connection)
                df.reset_index()
                df.to_excel(excel_writer=path,sheet_name='Compras Reporte',engine='openpyxl',startcol=0, startrow=4, index=False)
                self.styling_report(path)
                messagebox.showinfo('Reporte', 'Reporte Generado con Exito!', parent=self)


    def styling_report(self, path_file:str = ''):
        border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        def set_width_to(sheet, start, stop, width):
            for col in get_column_interval(start, stop):
                sheet.column_dimensions[col].width = width
                for x in range(min_row+1, max_row+1):
                    sheet[f'{col}{x}'].border = border_style
                    alignment_style = Alignment(horizontal='center', vertical='center')
                    sheet[f'{col}{x}'].alignment = alignment_style

        currency_style = NamedStyle(name='currency_style',
             font=Font(size=10),
             alignment=Alignment(horizontal='center', vertical='center'),
             number_format=FORMAT_CURRENCY_USD_SIMPLE,
        )


        informe = openpyxl.load_workbook(path_file)

        sheet = informe.active 
        min_row = sheet.min_row
        max_row = sheet.max_row

        for row in sheet[f'L{min_row+1}:N{max_row}']:
            for cell in row:
                cell.style = currency_style

        row_statistic = max_row+3
        sheet[f'A{row_statistic}'] = 'Resumen'
        sheet[f'A{row_statistic}'].font = Font(name='Arial', size=11, bold=True)
        sheet[f'A{row_statistic}'].alignment = Alignment(horizontal='center', vertical='center')

        sheet.merge_cells(f'A{row_statistic}:C{row_statistic}')

        sheet[f'A{row_statistic+1}'] = 'Total Compras'
        sheet[f'B{row_statistic+1}'] = 'Total Deuda'
        sheet[f'C{row_statistic+1}'] = 'Total Pagado'

        sheet[f'A{row_statistic+1}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'B{row_statistic+1}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'C{row_statistic+1}'].font = Font(name='Arial', size=10, bold=True)

        sheet[f'A{row_statistic+1}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'B{row_statistic+1}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'C{row_statistic+1}'].alignment = Alignment(horizontal='center', vertical='center')

        sheet[f'A{row_statistic+2}'] = f'=SUM(L{min_row+1}:L{max_row})'
        sheet[f'B{row_statistic+2}'] = f'=SUM(M{min_row+1}:M{max_row})'
        sheet[f'C{row_statistic+2}'] = f'=SUM(N{min_row+1}:N{max_row})'

        for row in sheet[f'A{row_statistic+2}:C{row_statistic+2}']:
            for cell in row:
                cell.style = currency_style


        for row in sheet[f'A{row_statistic}:C{row_statistic+2}']:
            for cell in row:
                cell.border = border_style



        sheet['A1'] = 'Multiservicios Abielcaj91 C.A.'
        sheet['A1'].font = Font(name='Arial', size=12, bold=True)
        sheet['A2'] = 'Reporte de Compras '
        sheet['A2'].font = Font(name='Arial', size=10, bold=True)
        sheet['A3'] = 'Fecha'
        sheet['A3'].font = Font(name='Arial', size=10, bold=True)

        sheet['B3'] = datetime.today().strftime('%d/%m/%Y')
        sheet['B3'].font = Font(name='Arial', size=10)
        sheet.merge_cells('A1:C1')
        sheet.merge_cells('A2:C2')
        set_width_to(sheet, "A", "R", width=20)


        informe.save(path_file)
    
        
# app = ttb.Window(themename='new')
# SGDB_Style()
# Report(app, title='Facturas')
# app.mainloop()